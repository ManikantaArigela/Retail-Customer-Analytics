import os
import pandas as pd
from datetime import datetime
from sqlalchemy.orm import Session
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from reportlab.lib.pagesizes import letter
from reportlab.lib import colors
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from backend.services.analytics import get_user_dataframe, compute_kpis, get_product_performance
from backend.services.forecasting import generate_forecast_and_insights

EXPORTS_DIR = os.path.abspath(os.path.join(os.path.dirname(__file__), "..", "..", "data", "exports"))
os.makedirs(EXPORTS_DIR, exist_ok=True)

def get_user_exports_dir(user_id: int) -> str:
    user_dir = os.path.join(EXPORTS_DIR, f"user_{user_id}")
    os.makedirs(user_dir, exist_ok=True)
    return user_dir

def generate_csv_report(user_id: int, db: Session) -> str:
    """
    Generates a CSV export of the user's cleaned/processed sales records.
    """
    df = get_user_dataframe(user_id, db)
    if df.empty:
        raise ValueError("No sales data available to export.")
        
    export_path = os.path.join(get_user_exports_dir(user_id), "retail_data_export.csv")
    df.to_csv(export_path, index=False)
    return export_path

def generate_excel_report(user_id: int, db: Session) -> str:
    """
    Generates a professional multi-sheet Excel report.
    """
    df = get_user_dataframe(user_id, db)
    if df.empty:
        raise ValueError("No sales data available to export.")
        
    kpis = compute_kpis(df)
    insights = generate_forecast_and_insights(user_id, db)
    forecast = insights["forecast"]
    
    export_path = os.path.join(get_user_exports_dir(user_id), "retail_intelligence_report.xlsx")
    
    # Create workbook
    wb = Workbook()
    
    # ---------------------------------------------
    # Sheet 1: Executive Summary
    # ---------------------------------------------
    ws1 = wb.active
    ws1.title = "Executive Summary"
    ws1.views.sheetView[0].showGridLines = True
    
    # Styles
    title_font = Font(name="Calibri", size=16, bold=True, color="1F4E79")
    header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    label_font = Font(name="Calibri", size=11, bold=True)
    regular_font = Font(name="Calibri", size=11)
    
    header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    accent_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
    
    border_side = Side(border_style="thin", color="D3D3D3")
    thin_border = Border(left=border_side, right=border_side, top=border_side, bottom=border_side)
    
    # Content
    ws1["A1"] = "Retail Intelligence Platform - Executive Summary"
    ws1["A1"].font = title_font
    ws1["A2"] = f"Generated on: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
    ws1["A2"].font = Font(italic=True, size=10)
    
    # KPI Headers
    ws1.append([]) # A3 empty
    ws1.append(["Key Performance Indicators (KPIs)"]) # A4
    ws1["A4"].font = Font(size=12, bold=True, color="1F4E79")
    
    ws1.append(["Metric", "Value"]) # A5
    ws1["A5"].font = header_font
    ws1["B5"].font = header_font
    ws1["A5"].fill = header_fill
    ws1["B5"].fill = header_fill
    
    kpi_rows = [
        ("Total Revenue", kpis["total_revenue"]),
        ("Total Units Sold", kpis["total_sales"]),
        ("Total Customers", kpis["total_customers"]),
        ("Total Products Tracked", kpis["total_products"]),
        ("Total Profit", kpis["total_profit"]),
        ("Net Profit Margin (%)", f"{kpis['profit_margin']}%"),
        ("Average Stock Inventory Level", int(kpis["avg_inventory"])),
        ("Month-Over-Month Sales Growth (%)", f"{kpis['sales_growth']}%")
    ]
    
    for metric, val in kpi_rows:
        row_idx = ws1.max_row + 1
        ws1.cell(row=row_idx, column=1, value=metric).font = label_font
        cell = ws1.cell(row=row_idx, column=2, value=val)
        cell.font = regular_font
        if isinstance(val, (int, float)):
            cell.number_format = '₹#,##0.00' if 'Revenue' in metric or 'Profit' in metric else '#,##0'
        
        ws1.cell(row=row_idx, column=1).border = thin_border
        ws1.cell(row=row_idx, column=2).border = thin_border
        
    # Recommendations Table
    ws1.append([])
    ws1.append(["Strategic AI Recommendations"])
    ws1.cell(row=ws1.max_row, column=1).font = Font(size=12, bold=True, color="1F4E79")
    
    rec_header_idx = ws1.max_row + 1
    ws1.cell(row=rec_header_idx, column=1, value="AI Business Recommendation").font = header_font
    ws1.cell(row=rec_header_idx, column=1).fill = header_fill
    ws1.merge_cells(start_row=rec_header_idx, start_column=1, end_row=rec_header_idx, end_column=3)
    
    for rec in insights["recommendations"]:
        row_idx = ws1.max_row + 1
        cell = ws1.cell(row=row_idx, column=1, value=rec)
        cell.font = regular_font
        cell.fill = accent_fill
        cell.border = thin_border
        ws1.merge_cells(start_row=row_idx, start_column=1, end_row=row_idx, end_column=3)
        
    # Column width adjust
    ws1.column_dimensions['A'].width = 35
    ws1.column_dimensions['B'].width = 25
    ws1.column_dimensions['C'].width = 25
    
    # ---------------------------------------------
    # Sheet 2: Processed Sales Data
    # ---------------------------------------------
    ws2 = wb.create_sheet(title="Transaction Database")
    ws2.views.sheetView[0].showGridLines = True
    
    # Header row
    cols = list(df.columns)
    ws2.append(cols)
    for col_num, col_name in enumerate(cols, 1):
        cell = ws2.cell(row=1, column=col_num)
        cell.font = header_font
        cell.fill = header_fill
        
    for _, row in df.iterrows():
        # Clean datetime elements
        row_vals = []
        for col_name in cols:
            val = row[col_name]
            if isinstance(val, datetime):
                row_vals.append(val.strftime("%Y-%m-%d"))
            elif isinstance(val, pd.Timestamp):
                row_vals.append(val.strftime("%Y-%m-%d"))
            else:
                row_vals.append(val)
        ws2.append(row_vals)
        
    # Auto fit column widths
    for col in ws2.columns:
        max_len = max(len(str(cell.value or '')) for cell in col)
        col_letter = col[0].column_letter
        ws2.column_dimensions[col_letter].width = max(max_len + 3, 12)
        
    # ---------------------------------------------
    # Sheet 3: ML 30-Day Forecast
    # ---------------------------------------------
    ws3 = wb.create_sheet(title="30-Day ML Forecast")
    ws3.views.sheetView[0].showGridLines = True
    
    ws3.append(["Date", "Predicted Sales Revenue"])
    ws3["A1"].font = header_font
    ws3["B1"].font = header_font
    ws3["A1"].fill = header_fill
    ws3["B1"].fill = header_fill
    
    for point in forecast:
        row_idx = ws3.max_row + 1
        ws3.cell(row=row_idx, column=1, value=point["date"]).font = regular_font
        cell = ws3.cell(row=row_idx, column=2, value=point["predicted"])
        cell.font = regular_font
        cell.number_format = '₹#,##0.00'
        
        ws3.cell(row=row_idx, column=1).border = thin_border
        ws3.cell(row=row_idx, column=2).border = thin_border
        
    ws3.column_dimensions['A'].width = 15
    ws3.column_dimensions['B'].width = 25
    
    wb.save(export_path)
    return export_path

def generate_pdf_report(user_id: int, db: Session) -> str:
    """
    Generates a beautifully structured PDF Report for executives.
    """
    df = get_user_dataframe(user_id, db)
    if df.empty:
        raise ValueError("No sales data available to export.")
        
    kpis = compute_kpis(df)
    insights = generate_forecast_and_insights(user_id, db)
    top_products = get_product_performance(df, ascending=False, limit=5)
    
    export_path = os.path.join(get_user_exports_dir(user_id), "retail_intelligence_report.pdf")
    
    # PDF document template
    doc = SimpleDocTemplate(
        export_path,
        pagesize=letter,
        rightMargin=40,
        leftMargin=40,
        topMargin=40,
        bottomMargin=40
    )
    
    styles = getSampleStyleSheet()
    
    # Custom colors
    navy = colors.HexColor("#1F4E79")
    charcoal = colors.HexColor("#333333")
    light_blue = colors.HexColor("#E9EEF4")
    
    # Custom styles
    title_style = ParagraphStyle(
        'DocTitle',
        parent=styles['Heading1'],
        fontName='Helvetica-Bold',
        fontSize=24,
        textColor=navy,
        spaceAfter=6
    )
    
    subtitle_style = ParagraphStyle(
        'DocSub',
        parent=styles['Normal'],
        fontName='Helvetica-Oblique',
        fontSize=10,
        textColor=charcoal,
        spaceAfter=20
    )
    
    h2_style = ParagraphStyle(
        'H2',
        parent=styles['Heading2'],
        fontName='Helvetica-Bold',
        fontSize=14,
        textColor=navy,
        spaceBefore=15,
        spaceAfter=10
    )
    
    body_style = ParagraphStyle(
        'Body',
        parent=styles['BodyText'],
        fontName='Helvetica',
        fontSize=10,
        textColor=charcoal,
        leading=14
    )
    
    header_style = ParagraphStyle(
        'HeaderStyle',
        parent=styles['BodyText'],
        fontName='Helvetica-Bold',
        fontSize=10,
        textColor=colors.white,
        leading=14
    )
    
    rec_style = ParagraphStyle(
        'Recommendation',
        parent=styles['BodyText'],
        fontName='Helvetica-Bold',
        fontSize=9.5,
        textColor=navy,
        leading=14
    )
    
    story = []
    
    # Header Banner
    story.append(Paragraph("RETAIL INTELLIGENCE REPORT", title_style))
    story.append(Paragraph(f"Business Intelligence Summary — Generated on {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}", subtitle_style))
    story.append(Spacer(1, 10))
    
    # KPIs section
    story.append(Paragraph("Executive Performance Metrics", h2_style))
    
    kpi_data = [
        [
            Paragraph("<b>Total Revenue:</b>", body_style), Paragraph(f"₹{kpis['total_revenue']:,.2f}", body_style),
            Paragraph("<b>Total Products:</b>", body_style), Paragraph(f"{kpis['total_products']}", body_style)
        ],
        [
            Paragraph("<b>Total Units Sold:</b>", body_style), Paragraph(f"{kpis['total_sales']:,}", body_style),
            Paragraph("<b>Average Stock Level:</b>", body_style), Paragraph(f"{int(kpis['avg_inventory']):,}", body_style)
        ],
        [
            Paragraph("<b>Net Profit Margin:</b>", body_style), Paragraph(f"{kpis['profit_margin']:.2f}%", body_style),
            Paragraph("<b>Sales Growth (MoM):</b>", body_style), Paragraph(f"{kpis['sales_growth']:.2f}%", body_style)
        ]
    ]
    
    kpi_table = Table(kpi_data, colWidths=[120, 140, 120, 140])
    kpi_table.setStyle(TableStyle([
        ('BACKGROUND', (0,0), (-1,-1), light_blue),
        ('BOX', (0,0), (-1,-1), 1, navy),
        ('INNERGRID', (0,0), (-1,-1), 0.5, colors.white),
        ('TOPPADDING', (0,0), (-1,-1), 8),
        ('BOTTOMPADDING', (0,0), (-1,-1), 8),
        ('LEFTPADDING', (0,0), (-1,-1), 10),
        ('RIGHTPADDING', (0,0), (-1,-1), 10),
    ]))
    story.append(kpi_table)
    story.append(Spacer(1, 15))
    
    # Top Products
    story.append(Paragraph("Top Performing Products", h2_style))
    prod_headers = ["Product Name", "Units Sold", "Total Revenue", "Net Profit"]
    prod_data = [ [Paragraph(f"<b>{h}</b>", header_style) for h in prod_headers] ]
    
    for p in top_products:
        prod_data.append([
            Paragraph(p["product_name"], body_style),
            Paragraph(f"{p['quantity']:,}", body_style),
            Paragraph(f"₹{p['sales']:,.2f}", body_style),
            Paragraph(f"₹{p['profit']:,.2f}", body_style)
        ])
        
    prod_table = Table(prod_data, colWidths=[200, 90, 110, 120])
    prod_table.setStyle(TableStyle([
        ('BACKGROUND', (0,0), (-1,0), navy),
        ('BOTTOMPADDING', (0,0), (-1,-1), 6),
        ('TOPPADDING', (0,0), (-1,-1), 6),
        ('ROWBACKGROUNDS', (0,1), (-1,-1), [colors.white, light_blue]),
        ('GRID', (0,0), (-1,-1), 0.5, colors.lightgrey)
    ]))
    
    story.append(prod_table)
    story.append(Spacer(1, 15))
    
    # Strategic AI Recommendations
    story.append(Paragraph("Strategic AI Recommendations & Insights", h2_style))
    
    rec_bullets = []
    for rec in insights["recommendations"]:
        rec_bullets.append([Paragraph("•", rec_style), Paragraph(rec, body_style)])
        
    rec_table = Table(rec_bullets, colWidths=[20, 500])
    rec_table.setStyle(TableStyle([
        ('VALIGN', (0,0), (-1,-1), 'TOP'),
        ('BOTTOMPADDING', (0,0), (-1,-1), 8),
    ]))
    story.append(rec_table)
    
    doc.build(story)
    return export_path
